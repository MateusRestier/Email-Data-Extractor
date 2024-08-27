try:
    import imaplib
    import poplib
    import email
    from email.parser import Parser
    from email.header import decode_header
    import pandas as pd
    import os
    from datetime import datetime
    from email.utils import parsedate_to_datetime
    from openpyxl import load_workbook
    from datetime import timedelta
    import tkinter as tk
    import threading
    from openpyxl.styles import NamedStyle
    from dotenv import load_dotenv; load_dotenv() # Carregar as variáveis de ambiente do arquivo .env
except Exception:
    print(f"Erro ao importar as bibliotecas ou as variáveis de ambiente do arquivo .env. Verifique se as dependências estão corretas e tente novamente. Erro: {Exception}")


"""----------------------- INÍCIO DAS FUNÇÕES DE PROCESSAMENTO DE E-MAIL -----------------------"""


# Lista de remetentes que não devem ser marcados como correção de nota
remetentes_excluidos = [
    "ana.rebeca@bagaggio.com.br", "andrew.germano@bagaggio.com.br", "anny.kelly@bagaggio.com.br", 
    "crystiano.bento@bagaggio.com.br", "douglas.franca@bagaggio.com.br", "guilherme.almeida@bagaggio.com.br", 
    "juliana.almeida@bagaggio.com.br", "julio.correia@bagaggio.com.br", "matheus.silva@bagaggio.com.br", 
    "monique.barbosa@bagaggio.com.br", "pedro.rufino@bagaggio.com.br", "thamirys.abreu@bagaggio.com.br", 
    "vanessa.rebello@bagaggio.com.br", "centraldenotasbagaggio@outlook.com", "centraldenotas@bagaggio.com.br"#, "mateus.restier@bagaggio.com.br"
]


# Função principal para orquestrar o processo para um email específico
def process_email_account(email_user, email_pass, email_label, data_formatada, log_text_widget):
    # Tentar conexão via IMAP
    mail, protocol = connect_to_imap(email_user, email_pass, log_text_widget)
    if mail is None:
        # Se a conexão IMAP falhar, tentar via POP3
        mail, protocol = connect_to_pop(email_user, email_pass, log_text_widget)
        if mail is None:
            log(f"Falha ao conectar via IMAP e POP3 para {email_user}.")
            return []

    email_data = []

    if protocol == 'IMAP':
        message_ids = fetch_emails_imap(mail, data_formatada)
        for num in message_ids:
            email_info = process_email_imap(mail, num, email_label)  # Passar email_label aqui
            if email_info:
                email_info.append(email_label)  # Adicionar a origem do email
                email_data.append(email_info)
    elif protocol == 'POP3':
        num_emails = len(mail.list()[1])
        emails = fetch_emails_pop(mail, num_emails, data_formatada)
        for msg in emails:
            email_info = process_email(msg, email_label)  # Passar email_label aqui
            if email_info:
                email_info.append(email_label)  # Adicionar a origem do email
                email_data.append(email_info)

    if protocol == 'IMAP':
        mail.logout()
    else:
        mail.quit()

    return email_data


# Função para tentar conexão IMAP
def connect_to_imap(email_user, email_pass, log_text_widget):
    servidores_imap = [
        "outlook.office365.com",
        "imap-mail.outlook.com",
        "smtp-mail.outlook.com"
    ]
    for servidor in servidores_imap:
        log(f"\n\nTentando conexão IMAP com o servidor: {servidor} e email: {email_user}", log_text_widget)
        try:
            mail = imaplib.IMAP4_SSL(servidor)
            mail.login(email_user, email_pass)
            log(f"\nConectado com sucesso à conta {email_user} via IMAP no servidor {servidor}", log_text_widget)
            return mail, 'IMAP'
        except imaplib.IMAP4.error as e:
            log(f"Erro de autenticação IMAP com o servidor {servidor}: {e}", log_text_widget)
        except Exception as e:
            log(f"Erro desconhecido ao tentar o servidor IMAP {servidor}: {e}", log_text_widget)
    log("\nFalha ao conectar via IMAP.", log_text_widget)
    return None, None


# Função para tentar conexão POP3
def connect_to_pop(email_user, email_pass, log_text_widget):
    servidores_pop3 = [
        "smtp-mail.outlook.com",
        "outlook.office365.com",
        "pop-mail.outlook.com"
    ]
    for servidor in servidores_pop3:
        log(f"\n\n\nTentando conexão POP3 com o servidor: {servidor} e email: {email_user}", log_text_widget)
        try:
            mail = poplib.POP3_SSL(servidor)
            mail.user(email_user)
            mail.pass_(email_pass)
            log(f"\nConectado com sucesso à conta {email_user} via POP3 no servidor {servidor}", log_text_widget)
            return mail, 'POP3'
        except poplib.error_proto as e:
            log(f"Erro de autenticação POP3 com o servidor {servidor}: {e}", log_text_widget)
        except Exception as e:
            log(f"Erro desconhecido ao tentar o servidor POP3 {servidor}: {e}", log_text_widget)
    log("\nFalha ao conectar via POP3.\n")
    return None, None


# Função para buscar e-mails por data específica via IMAP
def fetch_emails_imap(mail, date):
    mail.select("inbox")
    date_obj = datetime.strptime(date, "%Y-%m-%d")
    imap_date = date_obj.strftime("%d-%b-%Y")
    status, messages = mail.search(None, f'(ON {imap_date})')
    return messages[0].split()


# Função para buscar e-mails por data específica via POP3
def fetch_emails_pop(mail, num_emails, target_date):
    email_data = []
    for i in range(num_emails):
        raw_email = b"\n".join(mail.retr(i + 1)[1]).decode('utf-8')
        msg = Parser().parsestr(raw_email)

        # Processar e verificar a data
        email_date = parsedate_to_datetime(msg['Date']).strftime('%Y-%m-%d')
        if email_date == target_date:
            email_data.append(msg)
    return email_data


# Função para processar um e-mail e extrair os dados para IMAP
def process_email_imap(mail, num, email_label):
    status, msg_data = mail.fetch(num, "(RFC822)")
    for response_part in msg_data:
        if isinstance(response_part, tuple):
            msg = email.message_from_bytes(response_part[1])
            return process_email(msg, email_label)  # Passar email_label aqui


# Função para processar um e-mail e extrair os dados (comum para ambos IMAP e POP3)
def process_email(msg, email_label):
    # Decodificar o assunto
    subject, encoding = decode_header(msg["Subject"])[0]
    if isinstance(subject, bytes):
        try:
            subject = subject.decode(encoding if encoding else "utf-8")
        except UnicodeDecodeError:
            subject = subject.decode(encoding if encoding else "windows-1252")

    # Decodificar o remetente
    sender, encoding = decode_header(msg["From"])[0]
    if isinstance(sender, bytes):
        try:
            sender = sender.decode(encoding if encoding else "utf-8")
        except UnicodeDecodeError:
            sender = sender.decode(encoding if encoding else "windows-1252")

    # Extraia apenas o e-mail do campo From (se houver nome junto com o e-mail)
    if '<' in sender and '>' in sender:
        sender = sender.split('<')[1].replace('>', '').strip()
    else:
        sender = sender.strip()

    # Identificar se o e-mail é uma correção de nota e atribuir a data recebida ao invés de "S"
    if sender not in remetentes_excluidos:
        # Usar a data formatada como 'YYYY-MM-DD'
        correcao_nota = parsedate_to_datetime(msg["Date"]).strftime('%Y-%m-%d')
    else:
        correcao_nota = ''

    # Processar a data do e-mail
    date_received = msg["Date"]

    # O resto do código permanece o mesmo
    attachment_count = 0
    attachment_extensions = []
    attachment_names = []
    nf_count = 0
    nf_files = []
    compressed_count = 0

    for part in msg.walk():
        if part.get_content_disposition() == 'attachment':
            attachment_count += 1
            filename = part.get_filename()

            # Verificar se o nome do arquivo é válido
            if filename:
                # Decodificar o nome do arquivo se necessário
                decoded_filename, encoding = decode_header(filename)[0]
                if isinstance(decoded_filename, bytes):
                    try:
                        filename = decoded_filename.decode(encoding if encoding else 'utf-8')
                    except UnicodeDecodeError:
                        filename = decoded_filename.decode(encoding if encoding else 'windows-1252')

                attachment_names.append(filename)

                # Extrair a extensão do arquivo e corrigir se houver malformação
                extension = os.path.splitext(filename)[1][1:].lower()

                if extension.endswith('?='):
                    extension = extension.replace('?=', '')

                attachment_extensions.append(extension)

                if "nf" in filename.lower():
                    nf_count += 1
                    nf_files.append(filename)

                if extension in ['zip', 'rar']:
                    compressed_count += 1
            else:
                log(f"Anexo encontrado, mas sem nome de arquivo.")
                attachment_names.append("Sem nome")
                attachment_extensions.append("desconhecido")

    extensions_str = ','.join(attachment_extensions) if attachment_extensions else 'Nenhum'
    attachment_names_str = ' | '.join(attachment_names) if attachment_names else 'Nenhum'
    nf_files_str = ', '.join(nf_files) if nf_files else 'Nenhum'

    return [subject, date_received, sender, attachment_count, nf_count, extensions_str, compressed_count, attachment_names_str, nf_files_str, correcao_nota]


"""----------------------- INÍCIO DAS FUNÇÕES DE TRATAMENTO DE DADOS -----------------------"""


# Função para formatar a data
def formatar_data(data):
    try:
        data_formatada = datetime.strptime(data, "%Y%m%d").strftime("%Y-%m-%d")
        return data_formatada
    except ValueError:
        return None


# Função para definir larguras fixas para as colunas
def set_fixed_column_widths(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    # Definir larguras fixas para cada coluna
    column_widths = {
        "A": 50,  # Assunto
        "B": 12,  # Data
        "C": 16,  # Data Devolução
        "D": 16,  # Correção de Nota
        "E": 45,  # Remetente
        "F": 8,   # Anexos
        "G": 12,  # Notas Fiscais
        "H": 16,  # Extensões
        "I": 20,  # Arquivos Comprimidos
        "J": 45,  # Nomes dos Arquivos
        "K": 30,  # Arquivos NF
        "L": 18,  # Origem do Email
    }

    # Aplicar as larguras para cada coluna
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(excel_path)


# Função para mover a coluna "Correção de Nota" para a posição ao lado da coluna "Data Devolução"
def move_correction_column(df):
    # Obter as colunas atuais em uma lista
    cols = df.columns.tolist()

    # Remover a coluna "Correção de Nota" da sua posição atual
    cols.remove('Correção de Nota')

    # Inserir "Correção de Nota" na posição 3 (logo após "Data Devolução")
    cols.insert(3, 'Correção de Nota')

    # Reorganizar o DataFrame com a nova ordem de colunas
    df = df[cols]
    
    return df


# Função para converter as datas no formato 'Fri, 16 Aug 2024 15:04:39 +0000' para 'YYYY-MM-DD'
def convert_excel_dates(output_file_path):
    # Ler o arquivo Excel
    df = pd.read_excel(output_file_path)
    
    # Função para verificar e processar datas
    def parse_date(date_string):
        if pd.isna(date_string) or date_string == "":
            return date_string  # Se for NaN ou string vazia, retorna como está
        try:
            # Parse da string de data no formato 'Fri, 16 Aug 2024 15:04:39 +0000'
            parsed_date = datetime.strptime(date_string, "%a, %d %b %Y %H:%M:%S %z")
            # Retornar a data no formato 'YYYY-MM-DD'
            return parsed_date.strftime("%Y-%m-%d")
        except Exception as e:
            log(f"Erro ao processar a data {date_string}: {e}")
            return date_string  # Retorna a data original se houver erro

    # Aplicar a conversão na coluna de data, ignorando valores NaN ou inválidos
    df['Data'] = df['Data'].apply(parse_date)
    df['Data Devolução'] = df['Data Devolução'].apply(parse_date)
    
    # Salvar as alterações de volta no arquivo Excel
    df.to_excel(output_file_path, index=False)
    log(f"Datas convertidas e arquivo salvo com sucesso em {output_file_path}!")


# Função para definir o estilo de data no arquivo Excel
def set_date_format(excel_path):
    # Abrir o arquivo Excel existente
    wb = load_workbook(excel_path)
    ws = wb.active

    # Criar um estilo de data com o formato DD/MM/YYYY
    date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")

    # Aplicar o estilo de data às colunas de data (assumindo que Data está na coluna B, Data Devolução na C e Correção de Nota na D)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=4):  # B até D
        for cell in row:
            if cell.value is not None:
                cell.style = date_style

    # Salvar as alterações
    wb.save(excel_path)
    wb.close()


# Função para criar pastas dinâmicas para ano e mês
def create_dynamic_folders(base_dir, data_formatada, log_text_widget):
    """
    Cria pastas dinâmicas para ano e mês com base na data fornecida.

    :param base_dir: Diretório base onde as pastas serão criadas.
    :param data_formatada: Data formatada como 'YYYY-MM-DD'.
    :return: Caminho final da pasta onde o arquivo deve ser salvo.
    """
    # Obter o ano e o mês da data formatada
    data_obj = datetime.strptime(data_formatada, "%Y-%m-%d")
    ano = data_obj.strftime("%Y")
    mes = data_obj.strftime("%B") # Nome do mês completo

    # Criar a pasta do ano
    ano_folder = os.path.join(base_dir, ano)
    if not os.path.exists(ano_folder):
        os.makedirs(ano_folder)
        log(f"Pasta do ano '{ano}' criada.", log_text_widget)

    # Criar a pasta do mês dentro da pasta do ano
    mes_folder = os.path.join(ano_folder, mes)
    if not os.path.exists(mes_folder):
        os.makedirs(mes_folder)
        log(f"Pasta do mês '{mes}' criada dentro da pasta '{ano}'.", log_text_widget)

    # Retornar o caminho da pasta final onde o arquivo será salvo
    return mes_folder


# Função para salvar os dados no Excel
def save_to_excel(email_data, output_file_path):
    # Criar o DataFrame com as colunas atuais, incluindo a nova coluna "Correção de Nota"
    df = pd.DataFrame(email_data, columns=[
        "Assunto", "Data", "Remetente", "Anexos", "Notas Fiscais", "Extensões", 
        "Arquivos Comprimidos", "Nomes dos Arquivos", "Arquivos NF", "Correção de Nota", "Origem do Email"
    ])
    
    # Adicionar a coluna "Data Devolução" como uma cópia da coluna "Data"
    df.insert(2, "Data Devolução", df["Data"])  # Insere a nova coluna logo após "Data"
    
    # Esvaziar a coluna "Correção de Nota" para as linhas com "Central de Notas"
    df.loc[df['Origem do Email'] == 'Central de Notas', 'Correção de Nota'] = ''
    
    # Esvaziar a coluna "Data Devolução" para as linhas com "Central de Notas"
    df.loc[df['Origem do Email'] == 'Central de Notas', 'Data Devolução'] = ''
    
    # Salvar o DataFrame no arquivo Excel
    df.to_excel(output_file_path, index=False)
    log(f"\n\nInformações dos e-mails salvas com sucesso em: {output_file_path}!")


# Função para consolidar arquivos Excel
def consolidar_arquivos_excel(folder_path, data_formatada, log_text_widget):
    """
    Consolida todos os arquivos Excel na pasta fornecida e salva como um novo arquivo consolidado.
    
    :param folder_path: Caminho da pasta onde os arquivos Excel estão localizados.
    :param data_formatada: Data atual formatada para ser usada no nome do arquivo consolidado.
    """
    # Nome do arquivo consolidado
    consolidated_filename = f"Mapeamento_Notas_Fiscais_Consolidado_{data_formatada}.xlsx"
    consolidated_filepath = os.path.join(folder_path, consolidated_filename)

    # Verificar se já existem arquivos com "Mapeamento_Notas_Fiscais_Consolidado" no nome
    for file in os.listdir(folder_path):
        if file.startswith("Mapeamento_Notas_Fiscais_Consolidado"):
            try:
                os.remove(os.path.join(folder_path, file))
                log(f"\n\n\nArquivo consolidado antigo '{file}' removido com sucesso.", log_text_widget)
            except Exception as e:
                log(f"\n\n\nNão foi possível remover o arquivo '{file}': {e}. Ignorando e seguindo para o próximo.", log_text_widget)

    # Consolidar todos os arquivos Excel (exceto os consolidados)
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') and "Consolidado" not in f]
    
    # Lista para armazenar os DataFrames
    dfs = []
    
    log(f"\n\nOs arquivos abaixo, serão adicionados ao Consolidado:\n", log_text_widget)
    for excel_file in excel_files:
        file_path = os.path.join(folder_path, excel_file)
        try:
            df = pd.read_excel(file_path)
            dfs.append(df)
            log(f"{excel_file}.", log_text_widget)
        except Exception as e:
            log(f"Erro ao ler o arquivo '{excel_file}': {e}. Ignorando este arquivo.", log_text_widget)
    log("\n", log_text_widget)
    
    if dfs:
        # Concatenar todos os DataFrames
        consolidated_df = pd.concat(dfs, ignore_index=True)
        
        # Salvar o DataFrame consolidado como um novo arquivo Excel
        consolidated_df.to_excel(consolidated_filepath, index=False)
        log(f"\nArquivo consolidado salvo com sucesso em: {consolidated_filepath}", log_text_widget)
        
        # Ajustar as larguras das colunas do arquivo consolidado
        set_fixed_column_widths(consolidated_filepath)

        # Aplicar o formato de data no arquivo consolidado
        set_date_format(consolidated_filepath)  # Adicionar essa linha para aplicar o formato de data
    else:
        log("Nenhum arquivo Excel disponível para consolidação.", log_text_widget)


# Função para tratamento dos dados extraidos
def tratamento_dados(email_data, output_file_path, folder_path, data_formatada, log_text_widget):
    # Salvar os dados no Excel, adicionando a nova coluna 'Origem do Email'
    save_to_excel(email_data, output_file_path)

    # Converter datas e ajustar colunas
    convert_excel_dates(output_file_path)
    
    # Ler o arquivo Excel após as modificações
    df = pd.read_excel(output_file_path)

    # Mover a coluna "Correção de Nota" para a posição correta
    df = move_correction_column(df)

    # Converter 'Data Devolução', 'Data', e 'Correção de Nota' para datetime64
    df['Data Devolução'] = pd.to_datetime(df['Data Devolução'], errors='coerce')
    df['Data'] = pd.to_datetime(df['Data'], errors='coerce')
    df['Correção de Nota'] = pd.to_datetime(df['Correção de Nota'], errors='coerce')

    # Substituir por NaT ao invés de string vazia para valores nulos de data
    df.loc[df['Correção de Nota'].notna() & (df['Correção de Nota'] != ''), 'Data Devolução'] = pd.NaT

    # Salvar as alterações com a coluna movida
    df.to_excel(output_file_path, index=False)

    # Aplicar o formato de data no Excel
    set_date_format(output_file_path)

    # Definir larguras fixas para as colunas
    set_fixed_column_widths(output_file_path)
    
    # Consolidar arquivos Excel
    consolidar_arquivos_excel(folder_path, data_formatada, log_text_widget)


"""----------------------- INÍCIO DAS FUNÇÕES DA INTERFACE GRÁFICA -----------------------"""


# Função de log para o console
def log(message, log_widget=None):
    print(message)
    if log_widget:
        log_widget.insert(tk.END, message + '\n')
        log_widget.see(tk.END)


# Função para rodar o código em uma nova thread
def executar_em_thread(data_input, botao, status_label, log_text_widget):
    thread = threading.Thread(target=executar_codigo, args=(data_input, botao, status_label, log_text_widget))
    thread.start()


# Função que será executada ao clicar no botão ou pressionar Enter
def executar_codigo(data_input, botao, status_label, log_text_widget):  
    data_formatada = formatar_data(data_input)  # Formatando a data

    if data_formatada:
        status_label.config(text="Processando...", fg="white", bg='#053')
        botao.config(text="Pesquisando...", state="disabled")
        
        # Aqui você chama a função principal com a data formatada e o widget de log
        main(data_formatada, log_text_widget)
        
        status_label.config(text="Processo concluído!", fg="white", bg='#053')
        botao.config(text="Pesquisar", state="normal")
    else:
        status_label.config(text="Erro: Data inválida. Use o formato YYYYMMDD.", fg="red", bg='#053')


# Função para centralizar a janela na tela
def centralizar_janela(janela):
    largura_janela = 450
    altura_janela = 450

    largura_tela = janela.winfo_screenwidth()
    altura_tela = janela.winfo_screenheight()

    pos_x = (largura_tela // 2) - (largura_janela // 2)
    pos_y = (altura_tela // 2) - (altura_janela // 2)

    janela.geometry(f'{largura_janela}x{altura_janela}+{pos_x}+{pos_y}')


# Função para abrir a janela
def abrir_janela():
    # Criando a janela principal
    janela = tk.Tk()
    janela.title('Mapeamento de Notas Fiscais')
    janela.configure(bg='#053')

    # Centralizando a janela
    centralizar_janela(janela)

    # Texto de orientação
    texto_orientacao = tk.Label(janela, text="Insira a data (YYYYMMDD)", font=('Arial', 15, 'bold'), fg='white')
    texto_orientacao.configure(bg='#053')
    texto_orientacao.pack(padx=10, pady=20)

    # Obter dia, mês e ano atuais no formato YYYYMMDD (exemplo: '20240816')
    data_sugestao = datetime.now().strftime('%Y%m%d')

    # Criar o campo de entrada e inserir a sugestão com o dia atual
    global botao_abrir
    botao_abrir = tk.Entry(janela, font='Arial', width=20)
    botao_abrir.insert(0, data_sugestao)  # Sugere o dia atual
    botao_abrir.pack(pady=10)

    # Label de status
    global status_label
    status_label = tk.Label(janela, text="", font=('Arial', 12), fg="red", bg='#053')
    status_label.pack(pady=5)

    # Campo de texto para log
    log_text_widget = tk.Text(janela, height=10, width=50, wrap="word", bg='#053', fg='white')
    log_text_widget.pack(pady=10)

    # Botão de pesquisa
    global botao
    botao = tk.Button(janela, text="Pesquisar", font='Arial', width=20, command=lambda: executar_em_thread(botao_abrir.get(), botao, status_label, log_text_widget))
    botao.pack(padx=10, pady=40)

    # Iniciando a interface gráfica
    janela.mainloop()


"""----------------------- INÍCIO DA FUNÇÃO PRINCIPAL -----------------------"""


# Função principal para orquestrar o processo
def main(data_formatada, log_text_widget):
    log("\nＭａｐｅａｍｅｎｔｏ ｄｅ Ｎｏｔａｓ Ｆｉｓｃａｉｓ", log_text_widget)
    script_dir = os.getcwd()  # Pega o diretório atual onde o executável está sendo executado
    
    # Definir as credenciais dos dois emails
    email_user_1 = os.getenv('OUTLOOK_EMAILCDN')
    email_pass_1 = os.getenv('OUTLOOK_PASSWORDCDN')
    email_label_1 = "Central de Notas"

    email_user_2 = os.getenv('OUTLOOK_EMAILDCN')
    email_pass_2 = os.getenv('OUTLOOK_PASSWORDDCN')
    email_label_2 = "Devolução de Notas"

    # Processar os dois emails
    email_data_1 = process_email_account(email_user_1, email_pass_1, email_label_1, data_formatada, log_text_widget)
    email_data_2 = process_email_account(email_user_2, email_pass_2, email_label_2, data_formatada, log_text_widget)

    # Juntar os dados
    email_data = email_data_1 + email_data_2

    # Adicionar a nova coluna "Origem do Email"
    output_directory = os.path.join(script_dir, "Excel")  # Subpasta Excel dentro do mesmo diretório do script
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    # Criar pastas dinâmicas e obter o caminho final
    dynamic_folder_path = create_dynamic_folders(output_directory, data_formatada, log_text_widget)

    # Definir o caminho completo do arquivo com o diretório correto
    output_file_name = f"Mapeamento_Notas_Fiscais_{data_formatada}.xlsx"
    output_file_path = os.path.join(dynamic_folder_path, output_file_name)

    # Chamar a função de tratamento de dados
    tratamento_dados(email_data, output_file_path, dynamic_folder_path, data_formatada, log_text_widget)

    log(f"\nConexão ao servidor fechada. \nArquivo salvo em: {output_file_path}", log_text_widget)


if __name__ == "__main__":
    abrir_janela()